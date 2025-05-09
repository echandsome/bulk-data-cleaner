import os
import pandas as pd
import zipfile
import threading
from concurrent.futures import ProcessPoolExecutor, as_completed
from tkinter import *
from tkinter import filedialog, messagebox, ttk
import multiprocessing
import tempfile
import shutil
from datetime import datetime
from openpyxl import load_workbook
import csv

def excel_col_to_index(col):
    index = 0
    for i, char in enumerate(reversed(col)):
        index += (ord(char.upper()) - 64) * (26 ** i)
    return index - 1

EXCEL_INDEX = {
    'country': excel_col_to_index('W'),
    'language': excel_col_to_index('BE'),
    'occupation': excel_col_to_index('K'),
    'industry': excel_col_to_index('BG')
}

def process_group_external(country, records, output_dir):
    df = pd.DataFrame(records)
    sorted_df = df.sort_values(by=['Language', 'Occupation', 'Industry'])
    filename = os.path.join(output_dir, f"{country}.csv")
    sorted_df.to_csv(filename, index=False)

def filter_rachinbox_file(input_dir, filename, output_dir):
    try:
        path = os.path.join(input_dir, filename)
        df = pd.read_csv(path)
        df.columns = [f"Column{i+1}" for i in range(df.shape[1])]
        filtered_df = pd.DataFrame({
            "Email": df.iloc[:, 0],
            "First_Name": df.iloc[:, 2],
            "Last_Name": df.iloc[:, 3],
            "Company_Name": df.iloc[:, 14],
            "Linkdin": df.iloc[:, 13],
            "Personalised_Lines": df.iloc[:, 32]
        })
        filtered_filename = os.path.splitext(filename)[0] + "_rachInbox.csv"
        filtered_df.to_csv(os.path.join(output_dir, filtered_filename), index=False)
    except Exception as e:
        print(f"[rachInbox] Skipped {filename}: {str(e)}")

def filter_ghl_file(input_dir, filename, output_dir):
    try:
        path = os.path.join(input_dir, filename)
        df = pd.read_csv(path)
        df.columns = [f"Column{i+1}" for i in range(df.shape[1])]
        filtered_df = pd.DataFrame({
            "Email": df.iloc[:, 0],
            "First_Name": df.iloc[:, 2],
            "Last_Name": df.iloc[:, 3],
            "Department": df.iloc[:, 6],
            "Job_Title": df.iloc[:, 7],
            "Job_Level": df.iloc[:, 8],
            "City": df.iloc[:, 9],
            "State": df.iloc[:, 10],
            "Country": df.iloc[:, 11],
            "LinkedIn_Profile": df.iloc[:, 13],
            "Employer": df.iloc[:, 14],
            "Employer_Website": df.iloc[:, 15],
            "Phone": df.iloc[:, 16],
            "Employer_Facebook": df.iloc[:, 17],
            "Employer_LinkedIn": df.iloc[:, 18],
            "Employer_Founded_Date": df.iloc[:, 22],
            "Employer_Zip": df.iloc[:, 25],
            "Languages_Spoken": df.iloc[:, 28],
            "Industry": df.iloc[:, 29],
            "Focus": df.iloc[:, 30],
            "Skills": df.iloc[:, 31]
        })
        filtered_filename = os.path.splitext(filename)[0] + "_ghl.csv"
        filtered_df.to_csv(os.path.join(output_dir, filtered_filename), index=False)
    except Exception as e:
        print(f"[GHL] Skipped {filename}: {str(e)}")

def split_large_csv_files(src_folder, size_limit=48):

    for root, dirs, files in os.walk(src_folder):
        for file in files:
            if file.lower().endswith('.csv'):  # Only process CSV files
                file_path = os.path.join(root, file)
                file_size = os.path.getsize(file_path) / (1024 * 1024)  # Convert size to MB

                # Process file if it's larger than the specified limit
                if file_size >= size_limit:
                    print(f"Processing file {file} of size {file_size:.2f} MB")

                    with open(file_path, mode='r', newline='', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        header = next(reader)  # Read the header row
                        rows = list(reader)  # Read all the data rows

                    # Start splitting the file into smaller files
                    data_chunk = []
                    current_size = 0
                    part_number = 1

                    for row in rows:
                        # Add row to current chunk
                        data_chunk.append(row)
                        current_size += len(','.join(row))  # Approximate size by row length

                        # If current size exceeds the size limit, write to a new file
                        if current_size >= size_limit * 1024 * 1024:
                            # Define the new split file path
                            new_file_path = os.path.join(root, f"{file}_part_{part_number}.csv")
                            with open(new_file_path, mode='w', newline='', encoding='utf-8') as new_file:
                                writer = csv.writer(new_file)
                                writer.writerow(header)  # Write header to the new file
                                writer.writerows(data_chunk)  # Write the data chunk to the new file

                            print(f"Created: {new_file_path}")

                            # Reset for the next chunk
                            part_number += 1
                            data_chunk = []
                            current_size = 0

                    # If there are any remaining rows, write them to a final file
                    if data_chunk:
                        new_file_path = os.path.join(root, f"{file}_part_{part_number}.csv")
                        with open(new_file_path, mode='w', newline='', encoding='utf-8') as new_file:
                            writer = csv.writer(new_file)
                            writer.writerow(header)  # Write header to the new file
                            writer.writerows(data_chunk)  # Write the remaining data
                        print(f"Created: {new_file_path}")
                    
                    # Delete the large file
                    os.remove(file_path)
                    print(f"Deleted: {file} ({file_size:.2f} MB)")
                else:
                    print(f"File {file} is under {size_limit} MB, keeping it as is.")

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Bulk Processor")

        self.queue = []
        self.processing = False
        self.paused = False
        self.stopped = False
        self.save_dir = None
        self.temp_dirs = []

        self.setup_gui()

    def setup_gui(self):
        frm = Frame(self.root)
        frm.pack(padx=10, pady=10)

        self.file_label = Label(frm, text="No file selected")
        self.file_label.grid(row=0, column=0, columnspan=2)

        Button(frm, text="Select File", command=self.select_file).grid(row=1, column=0, columnspan=2, sticky='ew')
        Button(frm, text="Select Save Folder", command=self.select_save_folder).grid(row=2, column=0, columnspan=2, sticky='ew', pady=(10, 0))

        self.queue_box = Listbox(frm, height=8, width=50)
        self.queue_box.grid(row=3, column=0, columnspan=2, pady=5)

        self.start_stop_button = Button(frm, text="Start", command=self.toggle_start_stop)
        self.start_stop_button.grid(row=4, column=0, sticky='ew')

        self.pause_resume_button = Button(frm, text="Pause", command=self.toggle_pause_resume, state=DISABLED)
        self.pause_resume_button.grid(row=4, column=1, sticky='ew')

        self.status_label = Label(frm, text="Waiting...")
        self.status_label.grid(row=5, column=0, columnspan=2)

        self.progress = ttk.Progressbar(frm, orient=HORIZONTAL, length=400, mode='determinate')
        self.progress.grid(row=6, column=0, columnspan=2, pady=5)

    def toggle_start_stop(self):
        if not self.save_dir:
            messagebox.showwarning("Save Folder Required", "Please select a save folder before starting.")
            return
        if not self.processing:
            self.start_processing()
            self.start_stop_button.config(text="Stop")
            self.pause_resume_button.config(state=NORMAL)
        else:
            self.stop_processing()
            self.start_stop_button.config(text="Start")
            self.pause_resume_button.config(state=DISABLED)
            self.pause_resume_button.config(text="Pause")

    def toggle_pause_resume(self):
        if not self.paused:
            self.pause_processing()
            self.pause_resume_button.config(text="Resume")
        else:
            self.resume_processing()
            self.pause_resume_button.config(text="Pause")

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel or ZIP Files", "*.xlsx *.zip")])
        if not file_path:
            return

        if file_path.endswith('.xlsx'):
            self.queue.append(file_path)
            self.queue_box.insert(END, os.path.basename(file_path))
            self.file_label.config(text="No file selected")

        elif file_path.endswith('.zip'):
            temp_zip_dir = tempfile.mkdtemp()
            self.temp_dirs.append(temp_zip_dir)
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    xlsx_files = [f for f in zip_ref.namelist() if f.endswith('.xlsx')]
                    if not xlsx_files:
                        messagebox.showwarning("No Excel Files", "There are no .xlsx files in the ZIP file.")
                        return

                    for xlsx_file in xlsx_files:
                        extracted_path = zip_ref.extract(xlsx_file, temp_zip_dir)
                        self.queue.append(extracted_path)
                        self.queue_box.insert(END, os.path.basename(extracted_path))

                    self.file_label.config(text=f"{len(xlsx_files)} .xlsx files have been loaded from the ZIP.")
            except Exception as e:
                messagebox.showerror("ZIP Processing Error", str(e))

    def select_save_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.save_dir = folder_path
            self.status_label.config(text=f"Save folder selected: {folder_path}")

    def start_processing(self):
        if self.processing:
            return
        self.processing = True
        self.paused = False
        self.stopped = False
        threading.Thread(target=self.process_queue).start()

    def stop_processing(self):
        self.stopped = True
        self.processing = False
        self.queue.clear()
        self.queue_box.delete(0, END)
        self.status_label.config(text="Stopped")

    def pause_processing(self):
        self.paused = True
        self.status_label.config(text="Paused")

    def resume_processing(self):
        self.paused = False
        self.status_label.config(text="Resumed")

    def wait_if_paused_or_stopped(self):
        while self.paused and not self.stopped:
            self.root.update()
        if self.stopped:
            raise Exception("Processing stopped by user")

    def process_queue(self):
        while self.queue and not self.stopped:
            self.wait_if_paused_or_stopped()
            file = self.queue.pop(0)
            self.queue_box.delete(0)
            self.status_label.config(text=f"Processing {os.path.basename(file)}...")
            try:
                self.process_file(file)
            except Exception as e:
                self.status_label.config(text=str(e))
                break
            self.status_label.config(text=f"{os.path.basename(file)} completed")

        if self.stopped:
            self.status_label.config(text="All tasks cancelled")
        else:
            self.status_label.config(text="All tasks completed")
        self.start_stop_button.config(text="Start")
        self.pause_resume_button.config(state=DISABLED)
        self.pause_resume_button.config(text="Pause")
        self.processing = False

        # Clean up temporary directories
        for temp_dir in self.temp_dirs:
            shutil.rmtree(temp_dir, ignore_errors=True)
        self.temp_dirs.clear()

    def read_excel_with_progress(self, file_path):
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        data = []
        max_rows = ws.max_row

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            self.wait_if_paused_or_stopped()
            data.append(row)
            if i % 50 == 0 or i == max_rows:
                self.progress["value"] = int((i / max_rows) * 10)
                self.root.update_idletasks()

        wb.close()
        self.progress["value"] = 10
        return pd.DataFrame(data)

    def process_file(self, file_path):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.progress["value"] = 0
        self.root.update_idletasks()

        temp_dir = tempfile.mkdtemp()
        self.status_label.config(text="Reading Excel...")
        df = self.read_excel_with_progress(file_path)
        df.columns = [f"Column{i+1}" for i in range(df.shape[1])]

        try:
            rename_map = {
                df.columns[EXCEL_INDEX['country']]: 'Country',
                df.columns[EXCEL_INDEX['language']]: 'Language',
                df.columns[EXCEL_INDEX['occupation']]: 'Occupation',
                df.columns[EXCEL_INDEX['industry']]: 'Industry'
            }
            df = df.rename(columns=rename_map)
        except IndexError:
            self.status_label.config(text="Column index error")
            return

        df = self.clean_data(df)
        new_columns = [col if col in ['Country', 'Language', 'Occupation', 'Industry']
                       else f"Column{i+1}" for i, col in enumerate(df.columns)]
        df.columns = new_columns

        self.status_label.config(text="Processing by country...")
        grouped = df.groupby('Country')
        output_dir = os.path.join(os.path.dirname(temp_dir), "processed")
        os.makedirs(output_dir, exist_ok=True)

        country_futures = []
        with ProcessPoolExecutor(max_workers=max(1, multiprocessing.cpu_count() // 2)) as executor:
            for country, group in grouped:
                self.wait_if_paused_or_stopped()
                records = group.to_dict('records')
                future = executor.submit(process_group_external, country, records, output_dir)
                country_futures.append(future)

            total = len(country_futures)
            for i, f in enumerate(as_completed(country_futures), 1):
                self.wait_if_paused_or_stopped()
                self.progress["value"] = 10 + int((i / total) * 45)
                self.root.update_idletasks()

        self.zip_output(output_dir, os.path.dirname(file_path))

        self.status_label.config(text="Filtering rachInbox...")
        rach_dir = os.path.join(os.path.dirname(temp_dir), "rachInbox")
        self.filter_csvs_parallel(output_dir, rach_dir, os.path.dirname(file_path), filter_rachinbox_file, 55, 75)

        self.status_label.config(text="Filtering GHL...")
        ghl_dir = os.path.join(os.path.dirname(temp_dir), "ghl")
        self.filter_csvs_parallel(output_dir, ghl_dir, os.path.dirname(file_path), filter_ghl_file, 75, 95)

        self.progress["value"] = 100
        self.status_label.config(text="Processing complete.")

        shutil.rmtree(output_dir)
        shutil.rmtree(rach_dir)
        shutil.rmtree(ghl_dir)

    def filter_csvs_parallel(self, input_dir, output_dir, file_path, handler, start_pct, end_pct):
        os.makedirs(output_dir, exist_ok=True)
        files = [f for f in os.listdir(input_dir) if f.endswith(".csv")]
        futures = []

        with ProcessPoolExecutor() as executor:
            for file in files:
                self.wait_if_paused_or_stopped()
                futures.append(executor.submit(handler, input_dir, file, output_dir))

            total = len(futures)
            for i, f in enumerate(as_completed(futures), 1):
                self.wait_if_paused_or_stopped()
                self.progress["value"] = start_pct + int((i / total) * (end_pct - start_pct))
                self.root.update_idletasks()

        self.zip_output(output_dir, file_path)

    def clean_data(self, df):
        df = df.iloc[:, [i for i in range(df.shape[1]) if i % 2 == 0]]
        df = df.applymap(lambda x: '' if isinstance(x, str) and '#!$@-' in x else x)
        return df

    def zip_output(self, source_dir, default_destination_dir):
        split_large_csv_files(source_dir)
        folder_name = os.path.basename(source_dir.rstrip(os.sep))
        zip_filename = f"{folder_name}_{self.timestamp}.zip"
        zip_path = os.path.join(os.path.dirname(source_dir), zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in os.listdir(source_dir):
                full_path = os.path.join(source_dir, file)
                if os.path.isfile(full_path):
                    zipf.write(full_path, arcname=file)

        destination_dir = self.save_dir if self.save_dir else default_destination_dir
        os.makedirs(destination_dir, exist_ok=True)
        final_zip_path = os.path.join(destination_dir, zip_filename)
        shutil.move(zip_path, final_zip_path)

if __name__ == "__main__":
    multiprocessing.freeze_support()
    root = Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()
